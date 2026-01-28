import openpyxl
import os
import sys

# Force utf-8 for stdout if needed, but better to write to file
output_file = "inspection_result_utf8.txt"

template_path = "template/assessment.xlsx"
if not os.path.exists(template_path):
    print(f"Error: {template_path} not found.")
    exit()

wb = openpyxl.load_workbook(template_path)

with open(output_file, "w", encoding="utf-8") as f:
    f.write(f"Sheets: {wb.sheetnames}\n")

    # Inspect specific cells for diagnosis
    # Inspect specific keywords
    keywords = ["イラスト", "移動", "服薬", "インスリン", "就業"]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        f.write(f"\n--- Scanning Sheet: {sheet_name} ---\n")
        for row in ws.iter_rows(max_row=50): 
            for cell in row:
                 for kw in keywords:
                    if cell.value and isinstance(cell.value, str) and kw in cell.value:
                        f.write(f"Found '{kw}' in cell {cell.coordinate}: {cell.value}\n")

print(f"Done writing to {output_file}")
