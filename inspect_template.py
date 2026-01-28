import openpyxl
import os

template_path = "template/assessment.xlsx"
if not os.path.exists(template_path):
    print(f"Error: {template_path} not found.")
    exit()

wb = openpyxl.load_workbook(template_path)
print("Sheets:", wb.sheetnames)

target_text = "視線のあいにくさ"

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- Scanning Sheet: {sheet_name} ---")
    found = False
    for row in ws.iter_rows(max_row=50): # Scan top 50 rows
        for cell in row:
            if cell.value and isinstance(cell.value, str) and target_text in cell.value:
                print(f"Found '{target_text}' in cell {cell.coordinate}: {cell.value}")
                # Check neighbors to the right
                c = cell.column
                r = cell.row
                print(f"Neighboring cells (Right) for {cell.coordinate}:")
                for i in range(1, 10):
                    neighbor = ws.cell(row=r, column=c+i)
                    val = neighbor.value
                    if val:
                        print(f"  Offset +{i} ({neighbor.coordinate}): {val}")
                found = True
    if not found:
        print(f"'{target_text}' not found in {sheet_name}")
